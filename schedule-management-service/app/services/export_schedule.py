import io
import re
from datetime import date, time, datetime, timedelta
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session
from app.models import ScheduleSession, BreakConfig

class ExportScheduleService:

    @staticmethod
    def _clean_subject_name(name: str) -> str:
        """Removes content between parentheses from subject names"""
        if not name:
            return ""
        return re.sub(r'\(.*?\)', '', name).strip()

    @staticmethod
    def _get_headquarters(class_name: str) -> str:
        """Extracts headquarters based on the penultimate character of class name"""
        if not class_name or len(class_name) < 2:
            return "DESCONOCIDA"
            
        hq_code = class_name[-2].upper()
        hq_map = {
            '1': 'LIMA CERCADO',
            '2': 'SJL BASADRE',
            '3': 'INDEPENDENCIA',
            '4': '2 DE MAYO',
            '5': 'CONSTITUCION',
            '6': 'CRESPO Y CASTILLO',
            '8': 'SANTA ANITA',
            '9': 'COMAS',
            'Z': 'PUENTE PIEDRA',
            'Y': 'SJL LOS JARDINES',
            'X': 'VMT'
        }
        return hq_map.get(hq_code, 'OTRA')

    @staticmethod
    def _format_date_spanish(dt: date) -> str:
        """Formats date like 'lun 09 mar 2026'"""
        days = ['lun', 'mar', 'mié', 'jue', 'vie', 'sáb', 'dom']
        months = ['ene', 'feb', 'mar', 'abr', 'may', 'jun', 'jul', 'ago', 'sep', 'oct', 'nov', 'dic']
        
        day_str = days[dt.weekday()]
        month_str = months[dt.month - 1]
        
        return f"{day_str} {dt.day:02d} {month_str} {dt.year}"

    @staticmethod
    def _calculate_hours(t_start: time, t_end: time) -> float:
        """Calculates precise hours between two time objects"""
        dt1 = datetime.combine(date.today(), t_start)
        dt2 = datetime.combine(date.today(), t_end)
        diff = dt2 - dt1
        return round(diff.total_seconds() / 3600.0, 2)

    @staticmethod
    def _get_break_time(db: Session, t_start: time, t_end: time) -> float:
        """Calculates break overlap inside a given range in decimal hours based on break_config"""
        breaks = db.query(BreakConfig).all()
        total_break_seconds = 0
        
        dt_start = datetime.combine(date.today(), t_start)
        dt_end = datetime.combine(date.today(), t_end)
        
        for b in breaks:
            b_start = datetime.combine(date.today(), b.start_time)
            b_end = datetime.combine(date.today(), b.end_time)
            
            # Intersection logic
            overlap_start = max(dt_start, b_start)
            overlap_end = min(dt_end, b_end)
            
            if overlap_end > overlap_start:
                total_break_seconds += (overlap_end - overlap_start).total_seconds()
                
        return round(total_break_seconds / 3600.0, 2)

    @classmethod
    def generate_excel(cls, db: Session, sessions: List[ScheduleSession]) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Horarios Exportados"

        headers = [
            "Fecha de Clase", "Sede", "Ciclo", "Docente", 
            "Curso", "Hora Inicio", "Hora Fin", "Horas Dictadas", "Receso"
        ]

        # Header Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num = 2
        for s in sessions:
            teacher_name = f"{s.lesson.teacher.first_name} {s.lesson.teacher.last_name}" if s.lesson and s.lesson.teacher else ""
            course_name = cls._clean_subject_name(s.lesson.subject.name) if s.lesson and s.lesson.subject else ""
            class_name = s.lesson.class_group.name if s.lesson and s.lesson.class_group else ""
            
            sede = cls._get_headquarters(class_name)
            fecha_es = cls._format_date_spanish(s.session_date)
            horas_dictadas = cls._calculate_hours(s.start_time, s.end_time)
            receso_horas = cls._get_break_time(db, s.start_time, s.end_time)

            row_data = [
                fecha_es,
                sede,
                class_name,
                teacher_name,
                course_name,
                s.start_time.strftime("%H:%M"),
                s.end_time.strftime("%H:%M"),
                horas_dictadas,
                receso_horas
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                if col_num in [6, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal="center")

            row_num += 1

        # Autofit columns
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
