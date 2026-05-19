"""
Service del módulo DOCENTES — v2 completo.
Lógica de negocio: normalización, UPSERT Excel, cruce XML, reprocesamiento histórico + fuzzy.
"""
import logging
import unicodedata
import re
import time as time_lib
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional, Tuple

from sqlalchemy.orm import Session

import os
import json
from . import repository
from app.core.schemas import StandardResponse, PaginatedResponseData
from app.models import Teacher, XmlUpload, Lesson, ScheduleSession, TeacherNameOverride
from app.modules.observaciones import repository as obs_repo
from sqlalchemy import or_, distinct
from uuid import UUID

logger = logging.getLogger(__name__)

EVIDENCE_FILE = r"D:\Desktop\MOD HOR\schedule-management-service\mdm_evidence.json"

# --- CONSTANTES DE NORMALIZACIÓN v4 ---
ABBREVIATIONS = {
    "M.": "MARIA",
    "MZA.": "MARIA",
    "Mª": "MARIA",
    "FCO.": "FRANCISCO",
    "J.": "JOSE",
    "L.": "LUIS",
    "P.": "PEDRO",
    "R.": "RICARDO",
    "C.": "CARLOS",
    "A.": "ALBERTO",
    "FNDZ.": "FERNANDEZ",
    "GZ.": "GONZALEZ",
    "MTZ.": "MARTINEZ",
    "RZ.": "RODRIGUEZ",
}

PARTICLES = {"DE", "DEL", "LA", "LAS", "LO", "LOS", "Y"}
TITLES = {"PROF.", "LIC.", "DR.", "DRA.", "ING.", "MAG."}

def log_mdm_evidence(data: Dict[str, Any]):
    try:
        current_evidence = []
        if os.path.exists(EVIDENCE_FILE):
            with open(EVIDENCE_FILE, "r", encoding="utf-8") as f:
                current_evidence = json.load(f)
        
        from app.core.context import request_id_ctx
        current_evidence.append({
            **data,
            "request_id": request_id_ctx.get(),
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        
        with open(EVIDENCE_FILE, "w", encoding="utf-8") as f:
            json.dump(current_evidence, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logger.error(f"Error guardando evidencia MDM: {e}")


def normalize_name(text: str) -> str:
    """
    Normalización OBLIGATORIA (Fase de Cruce):
    - UPPERCASE
    - Sin acentos (á->a)
    - Ñ -> N
    - Sin espacios extra o dobles
    """
    if not text: return ""
    # 1. Uppercase
    t = text.upper()
    # 2. Quitar acentos y Ñ -> N
    nfkd = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in nfkd if not unicodedata.combining(c))
    # 3. Solo Alfanuméricos y espacios
    t = re.sub(r"[^A-Z0-9 ]", "", t)
    # 4. Quitar espacios dobles
    t = re.sub(r"\s+", " ", t).strip()
    return t


def normalize_for_match(text: str) -> str:
    """
    Normalización Estricta para Matching (v4.5):
    - Minúsculas, sin tildes, sin caracteres especiales, sin espacios dobles.
    """
    if not text: return ""
    # 1. Lowercase
    t = text.lower()
    # 2. Remove accents
    nfkd = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in nfkd if not unicodedata.combining(c))
    # 3. Remove special chars (keep alphanumeric and spaces)
    t = re.sub(r"[^a-z0-9 ]", "", t)
    # 4. Remove double spaces
    t = re.sub(r"\s+", " ", t).strip()
    return t



from app.core.config import settings
def resolve_teacher(db: Session, identifier: UUID) -> Teacher:
    """
    Resuelve un docente por ID (UUID) de manera estricta.
    Retorna el modelo Teacher o levanta un error si no lo encuentra.
    """
    if not identifier:
        raise ValueError("Identificador de docente no proporcionado")
    
    t = repository.fetch_teacher_by_id_full(db, identifier)
    if t:
        return t
        
    raise ValueError(f"No se encontró docente con el integrador UUID: {identifier}")


# ════════════════════════════════════════════════════════
#  NORMALIZACIÓN
# ════════════════════════════════════════════════════════

def ensure_string(value):
    if isinstance(value, dict):
        return value.get("canonical") or value.get("match") or ""

    if isinstance(value, str):
        return value.strip()

    logger.warning("normalized_name unexpected type", extra={"type": str(type(value)), "value": str(value)})
    return ""

def apply_crossing_result(teacher, result_type: str) -> None:
    if result_type == "CONFLICTO":
        new_status = "CONFLICTO"
    elif result_type == "SIN_ASIGNAR":
        new_status = "INCOMPLETO"
    else: # "MATCHED"
        if teacher.dni and teacher.dni.strip() and teacher.status in ["INCOMPLETO", "SIN_ASIGNAR"]:
            new_status = "ACTIVO"
        else:
            new_status = teacher.status

    if teacher.status != new_status:
        teacher.status = new_status

def normalize_teacher_name(apellidos: str, nombres: str) -> Dict[str, str]:
    """
    Normalización Dual (v4.0):
    1. canonical: [Paterno] [Materno] [Nombres] - Limpio pero legible.
    2. match: Versión sin partículas y tokens ordenados.
    """
    raw_ap = (apellidos or "").upper().strip()
    raw_nom = (nombres or "").upper().strip()
    
    all_text = f"{raw_ap} {raw_nom}"
    
    # Quitar tildes
    nfkd = unicodedata.normalize("NFKD", all_text)
    txt = "".join(c for c in nfkd if not unicodedata.combining(c))
    
    # Limpieza básica
    txt = re.sub(r"[,;.]", " ", txt) # Puntos se quitan para tokens pero ojo con M.
    # Re-evaluación: Si queremos expandir abrev, mejor no quitar puntos aún.
    
    # Mejor flujo:
    def preprocess(s: str) -> List[str]:
        # Expandir abreviaturas
        tokens = s.replace(",", " ").split()
        expanded = []
        for t in tokens:
            upper_t = t.upper()
            if upper_t in ABBREVIATIONS:
                expanded.append(ABBREVIATIONS[upper_t])
            elif upper_t.endswith(".") and upper_t in ABBREVIATIONS: # Doble chequeo
                 expanded.append(ABBREVIATIONS[upper_t])
            elif upper_t in TITLES:
                continue
            else:
                if len(upper_t) <= 2 and upper_t.endswith("."):
                    logger.warning(f"MDM_NORM: Abreviatura desconocida detectada: {upper_t}")
                expanded.append(upper_t)
        return expanded

    # 1. Canonical: Mantener estructura original limpia
    canonical_tokens = preprocess(f"{raw_ap} {raw_nom}")
    canonical = " ".join(canonical_tokens)
    
    # 2. Match Key: Quitar partículas y ordenar
    match_tokens = [t for t in canonical_tokens if t not in PARTICLES]
    match_tokens.sort()
    match_key = " ".join(match_tokens)
    
    return {
        "canonical": canonical.lower(),
        "match": match_key.lower()
    }


def normalize_single(name: str) -> str:
    """Fallback para compatibilidad legacy."""
    res = normalize_teacher_name(name, "")
    return res["canonical"]


def _fuzzy_ratio(a: str, b: str) -> float:
    """Ratio de similitud 0-100. Usa rapidfuzz si está instalado; fallback a difflib."""
    try:
        from rapidfuzz import fuzz
        return fuzz.ratio(a, b)
    except ImportError:
        from difflib import SequenceMatcher
        return SequenceMatcher(None, a, b).ratio() * 100

def determine_teacher_status(db: Session, last_name: str, first_name: str, dni: Optional[str], exclude_id: Optional[UUID] = None) -> str:
    """
    Gobernanza de Datos v5: Clasificación automática.
    """
    full_name = f"{last_name} {first_name}".upper().strip()
    
    # 1. INVALIDO (Basura o Pruebas)
    if len(full_name) < 4:
        return "INVALIDO"
    
    test_patterns = ["TEST", "PRUEBA", "DEMO", "ERROR", "BORRAR", "EJEMPLO", "PRUEBA", "PEREZ JOSE", "SANCHEZ MARIA"]
    for p in test_patterns:
        if p in full_name:
            return "INVALIDO"

    # 2. CONFLICTO (Duplicados detectados)
    if dni:
        existing_dni = repository.fetch_teacher_by_dni(db, dni)
        if existing_dni and (not exclude_id or existing_dni.id != exclude_id):
            return "CONFLICTO"

    norm = normalize_teacher_name(last_name, first_name)["canonical"]
    existing_norm = repository.fetch_teacher_by_normalized(db, norm)
    if existing_norm and (not exclude_id or existing_norm.id != exclude_id):
        # Si ya existe uno con el mismo nombre normalizado, es conflicto
        return "CONFLICTO"

    # 3. INCOMPLETO
    if not dni or not str(dni).strip():
        return "INCOMPLETO"

    # 4. ACTIVO (Todo OK)
    return "ACTIVO"


def merge_teachers_logic(db: Session, primary_id: UUID, secondary_id: UUID):
    """
    Consolida dos docentes en uno, reasignando todas las relaciones.
    """
    primary = repository.fetch_teacher_by_id_full(db, primary_id)
    secondary = repository.fetch_teacher_by_id_full(db, secondary_id)
    
    if not primary or not secondary:
        raise ValueError("Docente no encontrado")

    logger.info("[MERGE] Fusionando secondary %s into primary %s", secondary.id, primary.id)

    # 1. Consolidar datos básicos
    if not primary.dni: primary.dni = secondary.dni
    if not primary.razon_social: primary.razon_social = secondary.razon_social

    # 2. Reasignar relaciones
    from app.models.schedule import Lesson, Observation
    
    # Lessons
    db.query(Lesson).filter(Lesson.teacher_id == secondary_id).update({"teacher_id": primary_id})
    # Observations (as main teacher)
    db.query(Observation).filter(Observation.teacher_id == secondary_id).update({"teacher_id": primary_id})
    # Observations (as replacement)
    db.query(Observation).filter(Observation.replacement_teacher_id == secondary_id).update({"replacement_teacher_id": primary_id})

    # 3. Recalcular status del primary
    primary.status = determine_teacher_status(db, primary.last_name, primary.first_name, primary.dni, exclude_id=primary.id)
    
    # 4. Eliminar el secundario
    db.delete(secondary)
    db.commit()
    
    return _teacher_to_dict(primary)


# ════════════════════════════════════════════════════════
#  FUNCIÓN EXISTENTE — INTOCABLE
# ════════════════════════════════════════════════════════

def get_active_teachers_list(db: Session) -> List[str]:
    """Aplica la lógica de negocio para agrupar y formatear los nombres."""
    try:
        active_teachers_query = repository.fetch_active_teachers(db)
        return _format_teacher_list(active_teachers_query)
    except Exception as e:
        raise RuntimeError(f"Error procesando docentes activos: {str(e)}")


def get_active_teachers_for_rpt(db: Session) -> List[Dict[str, Any]]:
    """
    RPT Planillas Module Business Logic:
    1. Filter valid teachers (status == "ACTIVO", dni IS NOT NULL, merged_into_id IS NULL).
    2. Cross with the latest COMPLETED XML upload.
    3. Match using normalized_name.
    4. Parse XML classes and accumulate hours per teacher.
    """
    try:
        # Step 1: Filter valid teachers
        from app.models import Teacher
        from app.services.xml_parser import XMLParserService
        from datetime import time, datetime
        
        db_teachers = db.query(Teacher).filter(
            Teacher.status == "ACTIVO",
            Teacher.merged_into_id.is_(None)
        ).all()
        
        # Step 2: Get the latest COMPLETED XML upload
        upload = get_latest_completed_upload(db)
        if not upload or not upload.storage_path or not os.path.exists(upload.storage_path):
            logger.warning("[RPT] No se encontró carga XML válida para procesar horas.")
            return []
            
        # Parse XML
        parser = XMLParserService()
        parsed_data = parser.parse_file(upload.storage_path)
        
        # Parse XML teachers & build normalized_name maps
        xml_teachers = parsed_data.get("teachers", [])
        xml_teacher_map = {} # normalized_name -> list of xml_teacher_ids
        for xt in xml_teachers:
            fn = xt.get("first_name") or ""
            ln = xt.get("last_name") or ""
            # Calculate standard name from XML
            name = xt.get("name") or f"{ln} {fn}".strip()
            if not name:
                continue
            norm = normalize_name(name)
            if norm:
                if norm not in xml_teacher_map:
                    xml_teacher_map[norm] = []
                xml_teacher_map[norm].append(xt.get("source_id"))
                
        # Parse XML periods
        periods_map = parsed_data.get("periods", {})
        
        # Parse XML lessons & cards to accumulate hours per XML teacher source_id
        xml_lessons = parsed_data.get("lessons", [])
        lesson_teacher_map = {l.get("source_id"): l.get("teacher_id") for l in xml_lessons if l.get("source_id")}
        
        xml_teacher_hours = {} # xml_teacher_id -> total_hours
        cards = parsed_data.get("cards", [])
        for card in cards:
            lesson_id = card.get("lesson_id")
            xml_teacher_id = lesson_teacher_map.get(lesson_id)
            if not xml_teacher_id:
                continue
                
            period_id = card.get("period")
            if period_id is None:
                continue
            try:
                period_id = int(period_id)
            except ValueError:
                continue
                
            period_time = periods_map.get(period_id)
            if not period_time:
                continue
                
            start_time = period_time.get("start")
            end_time = period_time.get("end")
            if not start_time or not end_time:
                continue
                
            # Calculate duration in minutes
            def parse_time_str(t):
                if isinstance(t, time):
                    return t
                for fmt in ("%H:%M:%S", "%H:%M"):
                    try:
                        return datetime.strptime(t, fmt).time()
                    except ValueError:
                        continue
                return None
                
            t_start = parse_time_str(start_time)
            t_end = parse_time_str(end_time)
            if not t_start or not t_end:
                continue
                
            duration_minutes = (t_end.hour * 60 + t_end.minute) - (t_start.hour * 60 + t_start.minute)
            # academic hours = duration_minutes / 50.0
            academic_hours = duration_minutes / 50.0
            
            # Days count
            days_str = card.get("days", "1")
            days_count = days_str.count("1") if days_str else 1
            
            card_hours = academic_hours * days_count
            xml_teacher_hours[xml_teacher_id] = xml_teacher_hours.get(xml_teacher_id, 0.0) + card_hours
            
        # Match valid DB teachers with XML teachers using normalized_name
        result = []
        for teacher in db_teachers:
            db_norm = normalize_name(f"{teacher.last_name} {teacher.first_name}")
            matched_xml_ids = xml_teacher_map.get(db_norm)
            if not matched_xml_ids:
                # Try sorting tokens for more flexible matching
                sorted_tokens_db = " ".join(sorted(db_norm.split()))
                for xml_norm, xml_ids in xml_teacher_map.items():
                    sorted_tokens_xml = " ".join(sorted(xml_norm.split()))
                    if sorted_tokens_db == sorted_tokens_xml:
                        matched_xml_ids = xml_ids
                        break
                        
            if matched_xml_ids:
                if not teacher.dni:
                    logger.info(f"[TEACHER NULL DNI ACCEPTED] Teacher ID {teacher.id} allowed without DNI.")
                
                # Accumulate hours from all matching XML teacher IDs (if any)
                total_hours = sum(xml_teacher_hours.get(xt_id, 0.0) for xt_id in matched_xml_ids)
                result.append({
                    "teacher_id": teacher.id,
                    "name": f"{teacher.last_name}, {teacher.first_name}".strip().upper(),
                    "dni": teacher.dni,
                    "total_hours": round(total_hours, 2)
                })
                
        # Sort by name alphabetically
        result.sort(key=lambda x: x["name"])
        logger.info(f"[API PAYLOAD NORMALIZED] successfully returning {len(result)} teacher hours items.")
        return result
        
    except Exception as e:
        logger.error(f"Error calculating RPT teachers and hours: {str(e)}", exc_info=True)
        raise RuntimeError(f"Error calculating RPT teachers and hours: {str(e)}")


def _format_teacher_list(teachers_query) -> List[str]:
    unique_names = set()
    grouped_by_last_name = {}

    for t in teachers_query:
        fn = (t.first_name or "").strip().upper()
        ln = (t.last_name or "").strip().upper()
        formatted = f"{ln}, {fn}"
        unique_names.add(formatted)

    for name in unique_names:
        parts = name.split(",")
        if len(parts) >= 2:
            last_names = parts[0].strip()
            first_names = parts[1].strip()
            if last_names not in grouped_by_last_name:
                grouped_by_last_name[last_names] = name
            else:
                if len(first_names) > len(grouped_by_last_name[last_names].split(",")[1].strip()):
                    grouped_by_last_name[last_names] = name
        else:
            grouped_by_last_name[name] = name

    final_list = list(grouped_by_last_name.values())
    final_list.sort()
    return final_list


# ════════════════════════════════════════════════════════
#  MAESTRA — CRUD DE TEACHERS
# ════════════════════════════════════════════════════════

def _teacher_to_dict(t: Teacher, active_ids: set = None) -> Dict[str, Any]:
    # Detectamos si 'id' es UUID o legacy
    t_id_str = str(t.id)
    legacy_id = getattr(t, 'legacy_id', None)

    return {
        "id": t_id_str, # Siempre UUID en la nueva arquitectura
        "legacy_id": legacy_id,
        "source_id": t.source_id,
        "nombres": t.first_name or "",
        "apellidos": t.last_name or "",
        "short_name": t.short_name or "",
        "dni": t.dni or "",
        "razon_social": t.razon_social or "",
        "normalized_name": t.normalized_name or "",
        "is_active": t.is_active,
        "is_assigned": getattr(t, 'is_assigned', True),
        "possible_duplicate": getattr(t, 'possible_duplicate', False),
        "times_detected": getattr(t, 'times_detected', 0),
        "last_seen_at": t.last_seen_at.strftime("%Y-%m-%d") if getattr(t, 'last_seen_at', None) else "",
        "is_active_by_workload": (t.id in active_ids) if active_ids is not None else False,
    }


def get_all_teachers(
    db: Session,
    filter_mode: str = "all",
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    status_filter: Optional[str] = None,
):
    result = repository.fetch_all_teachers_paginated(db, filter_mode, search, page, limit, status_filter)
    active_ids = {t.id for t in repository.fetch_active_teachers(db)}
    
    data_items = [_teacher_to_dict(t, active_ids) for t in result["items"]]
    
    return StandardResponse(
        success=True,
        data=PaginatedResponseData(
            data=data_items,
            total=result["total"],
            page=result["page"],
            limit=limit,
            total_pages=result["total_pages"]
        )
    )


def create_teacher_manual(db: Session, payload: Dict[str, Any]) -> Dict[str, Any]:
    apellidos = (payload.get("last_name") or "").strip()
    nombres   = (payload.get("first_name") or "").strip()
    dni       = (payload.get("dni") or "").strip() or None
    norm_dict = normalize_teacher_name(apellidos, nombres)
    norm      = ensure_string(norm_dict)

    if dni:
        existing = repository.fetch_teacher_by_dni(db, dni)
        if existing:
            raise ValueError(f"Ya existe un docente con DNI {dni} (id={existing.id})")

    existing = repository.fetch_teacher_by_normalized(db, norm)
    if existing:
        raise ValueError(f"Ya existe un docente similar: {existing.last_name} {existing.first_name}")

    status = determine_teacher_status(db, apellidos, nombres, dni)
    data = {
        "source_id":       f"MANUAL_{int(time_lib.time())}",
        "first_name":      nombres,
        "last_name":       apellidos,
        "short_name":      payload.get("short_name", ""),
        "dni":             dni,
        "razon_social":    (payload.get("razon_social") or "").strip() or None,
        "normalized_name": norm,
        "status":          status,
    }
    t = repository.create_teacher(db, data)
    apply_crossing_result(t, "MATCHED")
    db.commit()
    return _teacher_to_dict(t)


def update_teacher_data(db: Session, teacher_id: UUID, payload: Dict[str, Any]) -> Dict[str, Any]:
    t = resolve_teacher(db, teacher_id)
    if not t:
        raise ValueError("Docente no encontrado")

    # Concurrency Protection
    expected_updated_at = payload.get("updated_at")
    if expected_updated_at and getattr(t, "updated_at", None):
        if str(t.updated_at) != str(expected_updated_at):
            raise ValueError("El registro ha sido modificado por otro usuario. Por favor recargue la página.")

    apellidos = (payload.get("last_name") or t.last_name or "").strip()
    nombres   = (payload.get("first_name") or t.first_name or "").strip()
    norm_dict = normalize_teacher_name(apellidos, nombres)
    norm      = ensure_string(norm_dict)

    dni = (payload.get("dni") or t.dni or "").strip() or None
    status = determine_teacher_status(db, apellidos, nombres, dni, exclude_id=t.id)

    updated = repository.update_teacher(db, t, {
        "first_name":      nombres,
        "last_name":       apellidos,
        "short_name":      payload.get("short_name", t.short_name or ""),
        "dni":             dni,
        "razon_social":    (payload.get("razon_social") or "").strip() or None,
        "normalized_name": norm,
        "status":          status,
    })
    
    apply_crossing_result(updated, "MATCHED")
    db.commit()
    return _teacher_to_dict(updated)


def update_teacher_status(db: Session, teacher_id: UUID, is_active: bool) -> Dict[str, Any]:
    """Actualiza solo el campo is_active."""
    t = resolve_teacher(db, teacher_id)
    if not t:
        raise ValueError("Docente no encontrado")
    
    updated = repository.update_teacher(db, t, {"is_active": is_active})
    db.commit()
    return _teacher_to_dict(updated)


def get_teacher_activity_info(db: Session, teacher_id: UUID) -> Dict[str, bool]:
    """Retorna si el docente tiene actividad para el modal de advertencia v3.10."""
    t = resolve_teacher(db, teacher_id)
    return repository.check_teacher_activity(db, t.id)


# ════════════════════════════════════════════════════════
#  EXCEL IMPORT (openpyxl, sin pandas)
# ════════════════════════════════════════════════════════

def import_excel(db: Session, file_bytes: bytes) -> Dict[str, Any]:
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    headers_raw = [str(ws.cell(1, c).value or "").strip().upper() for c in range(1, ws.max_column + 1)]
    logger.info("[IMPORT] max_row=%s  headers=%s", ws.max_row, headers_raw)

    col_map: Dict[str, int] = {}
    ALIASES = {
        "DNI":          ["DNI", "RUC", "DOCUMENTO"],
        "APELLIDOS":    ["APELLIDOS", "APELLIDO", "LAST_NAME"],
        "NOMBRES":      ["NOMBRES", "NOMBRE", "FIRST_NAME"],
        "RAZON_SOCIAL": ["RAZON SOCIAL", "RAZÓN SOCIAL", "RAZON_SOCIAL", "EMPRESA"],
    }
    for canonical, aliases in ALIASES.items():
        for idx, h in enumerate(headers_raw, start=1):
            if h in aliases:
                col_map[canonical] = idx
                break

    logger.info("[IMPORT] col_map=%s", col_map)

    if "APELLIDOS" not in col_map or "NOMBRES" not in col_map:
        logger.error("[IMPORT] FAIL — faltan columnas obligatorias. headers=%s", headers_raw)
        raise ValueError(
            f"El archivo debe tener columnas APELLIDOS y NOMBRES como minimo. "
            f"Columnas detectadas: {headers_raw}"
        )

    rows_preview, inserted, updated, skipped = [], 0, 0, 0

    try:
        for row_idx in range(2, ws.max_row + 1):
            def cell(key):
                idx = col_map.get(key)
                if idx is None:
                    return ""
                val = ws.cell(row_idx, idx).value
                return str(val).strip() if val is not None else ""

            apellidos    = cell("APELLIDOS")
            nombres      = cell("NOMBRES")
            dni          = cell("DNI") or None
            razon_social = cell("RAZON_SOCIAL") or None

            if not apellidos and not nombres:
                skipped += 1
                logger.debug("[IMPORT] row=%d SKIP (vacia)", row_idx)
                continue

            norm_dict = normalize_teacher_name(apellidos, nombres)
            norm = norm_dict["canonical"]
            match_key = normalize_for_match(f"{apellidos} {nombres}")


            existing: Optional[Teacher] = None
            if dni:
                existing = repository.fetch_teacher_by_dni(db, dni)
            if not existing:
                existing = repository.fetch_teacher_by_normalized(db, norm)

            status = determine_teacher_status(db, apellidos, nombres, dni, exclude_id=existing.id if existing else None)
            print(f"[CLASSIFY] {apellidos} {nombres} -> {status}")

            if existing:
                repository.update_teacher(db, existing, {
                    "first_name":           nombres,
                    "last_name":            apellidos,
                    "dni":                  dni,
                    "razon_social":         razon_social,
                    "normalized_name":      norm,
                    "normalized_for_match": match_key,
                    "is_assigned":          True,
                    "is_active":            True,
                    "status":               status
                })
                updated += 1
                action = "updated"
                logger.info("[IMPORT] row=%d UPDATE: dni=%s ap=%s nom=%s status=%s", row_idx, dni, apellidos, nombres, status)
            else:
                repository.create_teacher(db, {
                    "source_id":            f"EXCEL_{int(time_lib.time())}_{row_idx}",
                    "first_name":           nombres,
                    "last_name":            apellidos,
                    "short_name":           "",
                    "dni":                  dni,
                    "razon_social":         razon_social,
                    "normalized_name":      norm,
                    "normalized_for_match": match_key,
                    "is_assigned":          True,   # ✅ FIX: Docentes Excel son de maestra
                    "is_active":            True,
                    "status":               status
                })
                inserted += 1
                action = "inserted"
                logger.info("[IMPORT] row=%d INSERT: dni=%s ap=%s nom=%s status=%s", row_idx, dni, apellidos, nombres, status)

            rows_preview.append({
                "fila":         row_idx,
                "apellidos":    apellidos,
                "nombres":      nombres,
                "dni":          dni or "",
                "razon_social": razon_social or "",
                "estado":       "INSERTADO" if action == "inserted" else "ACTUALIZADO",
                "mensaje":      f"Estado: {status}"
            })

        db.commit()
        logger.info(
            "[IMPORT] COMPLETADO: inserted=%d updated=%d skipped=%d total=%d",
            inserted, updated, skipped, inserted + updated + skipped
        )
        
        return StandardResponse(
            success=True,
            data={
                "data": rows_preview,
                "total": len(rows_preview),
                "page": 1,
                "limit": max(50, len(rows_preview)),
                "total_pages": 1,
                "inserted": inserted,
                "updated": updated,
                "skipped": skipped
            }
        )

    except Exception as e:
        db.rollback()
        logger.error("[IMPORT] ERROR: %s", str(e), exc_info=True)
        raise RuntimeError(f"Error importando Excel: {str(e)}")


def bulk_delete_excel_teachers(db: Session, executor_email: str) -> Dict[str, Any]:
    """
    Realiza el Reset Operativo Controlado de manera transaccional:
    1. Trunca la tabla observations.
    2. Elimina rpt_planilla y schedule_sessions vinculados al XML histórico.
    3. Elimina lecciones huérfanas en cascada.
    4. Elimina selectivamente docentes importados desde Excel.
    5. Elimina el registro del XML de la BD y su archivo físico de almacenamiento.
    """
    import uuid
    from datetime import datetime, timezone
    from sqlalchemy import text
    import os

    trace_id = str(uuid.uuid4())
    logger.info("[CLEANUP START] Iniciando Reset Operativo Controlado. Trace ID: %s | Ejecutor: %s", trace_id, executor_email)

    try:
        # Buscamos el XML histórico para obtener su ID y ruta
        xml_filename = "historical_xml_import_202603.xml"
        xml_row = db.execute(
            text("SELECT id, storage_path FROM xml_uploads WHERE filename = :filename"),
            {"filename": xml_filename}
        ).fetchone()

        xml_id = str(xml_row[0]) if xml_row else "8bc2c3a5-fa43-4cb2-8971-ebd07ccb5b84"
        physical_path = xml_row[1] if xml_row else "storage/xml_uploads/historical_xml_import_202603.xml"

        # 1. Eliminar observaciones
        deleted_obs = db.execute(text("DELETE FROM observations")).rowcount

        # 2. Eliminar rpt_planilla
        deleted_rpt = db.execute(
            text("DELETE FROM rpt_planilla WHERE xml_upload_id = :xml_id"),
            {"xml_id": xml_id}
        ).rowcount

        # 3. Eliminar schedule_sessions
        deleted_sessions = db.execute(
            text("DELETE FROM schedule_sessions WHERE xml_upload_id = :xml_id"),
            {"xml_id": xml_id}
        ).rowcount

        # 4. Eliminar lecciones huérfanas
        deleted_lessons = db.execute(
            text("DELETE FROM lessons WHERE id NOT IN (SELECT DISTINCT lesson_id FROM schedule_sessions)")
        ).rowcount

        # 5. Eliminar docentes importados desde Excel
        deleted_teachers = db.execute(
            text("DELETE FROM teachers WHERE source_id LIKE 'EXCEL_%' AND source = 'manual'")
        ).rowcount

        # 6. Eliminar xml_upload
        deleted_uploads = db.execute(
            text("DELETE FROM xml_uploads WHERE id = :xml_id"),
            {"xml_id": xml_id}
        ).rowcount

        # 7. Eliminar archivo físico en storage si existe
        file_removed = False
        if physical_path and os.path.exists(physical_path):
            try:
                os.remove(physical_path)
                file_removed = True
                logger.info("[CLEANUP] Archivo físico XML eliminado de storage: %s", physical_path)
            except Exception as fe:
                logger.error("[CLEANUP ERROR] No se pudo eliminar archivo físico: %s. Error: %s", physical_path, str(fe))

        db.commit()

        utc_now = datetime.now(timezone.utc).isoformat()
        logger.info(
            "[SUPERADMIN OPERATION SUCCESS] | User: %s | Action: BULK_DELETE_EXCEL_TEACHERS | "
            "Timestamp: %s | Deleted Count: %d | Trace ID: %s",
            executor_email,
            utc_now,
            deleted_teachers,
            trace_id
        )

        return {
            "success": True,
            "data": {
                "trace_id": trace_id,
                "timestamp": utc_now,
                "deleted_counts": {
                    "teachers": deleted_teachers,
                    "observations": deleted_obs,
                    "rpt_planilla": deleted_rpt,
                    "schedule_sessions": deleted_sessions,
                    "lessons": deleted_lessons,
                    "xml_uploads": deleted_uploads,
                    "physical_file_removed": file_removed
                }
            },
            "error": None
        }

    except Exception as e:
        db.rollback()
        logger.error("[CLEANUP ERROR] Error durante Reset Operativo Controlado. Trace ID: %s. Detalle: %s", trace_id, str(e), exc_info=True)
        raise RuntimeError(f"Error durante Reset Operativo: {str(e)}")


# ════════════════════════════════════════════════════════
#  LÓGICA DE CRUCE (DB vs ÚLTIMO XML)
# ════════════════════════════════════════════════════════


from lxml import etree
import difflib

def get_latest_completed_upload(db: Session) -> Optional[XmlUpload]:
    """Obtiene la carga XML exitosa más reciente (Source of Truth)."""
    upload = db.query(XmlUpload).filter(XmlUpload.status == 'COMPLETED').order_by(XmlUpload.created_at.desc()).first()
    if not upload:
        logger.warning("[CRUCE] No se encontró ninguna carga COMPLETED en la DB.")
    return upload

def get_docente_names_from_xml(upload: XmlUpload) -> List[str]:
    """
    Parsea el XML almacenado y extrae los nombres normalizados de los docentes.
    """
    if not upload or not upload.storage_path:
        logger.warning("[CRUCE] El upload no tiene storage_path.")
        return []

    if not os.path.exists(upload.storage_path):
        logger.error(f"[CRUCE] El archivo XML no existe en el disco: {upload.storage_path}")
        return []

    try:
        tree = etree.parse(upload.storage_path)
        root = tree.getroot()
        # En el formato aSc Horarios, los docentes están en <teachers>
        teachers = root.xpath(".//teacher")
        names = []
        for t in teachers:
            # asctt suele tener name o (first_name + last_name)
            # Probamos ambos por seguridad
            name = t.get("name") or ""
            if not name:
                fn = t.get("firstname") or t.get("first_name") or ""
                ln = t.get("lastname") or t.get("last_name") or ""
                name = f"{ln} {fn}".strip()
            
            if name:
                names.append(normalize_name(name))
        
        # Eliminar duplicados
        unique_names = list(set(names))
        
        logger.info(f"Total docentes XML: {len(unique_names)}")
        if unique_names:
            logger.info(f"Ejemplo XML: {unique_names[:3]}")
        
        return unique_names
    except Exception as e:
        logger.error(f"[CRUCE] Error parseando XML: {str(e)}")
        return []

def check_strict_match(name1: str, name2: str) -> bool:
    """
    Validación ESTRICTA de matching para evitar falsos positivos:
    1. Misma cantidad de tokens (JUAN PEREZ != JUAN CARLOS PEREZ)
    2. El conjunto de tokens debe ser exactamente igual (el orden no importa)
    """
    tokens1 = name1.split()
    tokens2 = name2.split()
    
    if len(tokens1) != len(tokens2):
        return False
        
    if set(tokens1) != set(tokens2):
        logger.info(f"STRICT RULE FAILED: Token sets differ. XML='{name1}' DB='{name2}'")
        return False
        
    return True

def _run_crossing_engine(db: Session):
    latest = get_latest_completed_upload(db)
    xml_names = get_docente_names_from_xml(latest) if latest else []
    
    if not xml_names:
        return {"matched": [], "sin_asignar": [], "conflictos": []}

    db_teachers = db.query(Teacher).filter(Teacher.merged_into_id.is_(None)).all()
    db_names = []
    db_data_map = {} 
    db_sorted_data_map = {}
    db_map = {}
    db_sorted_map = {}
    
    for t in db_teachers:
        full_name = normalize_name(f"{t.last_name} {t.first_name}")
        db_names.append(full_name)
        if full_name not in db_data_map: db_data_map[full_name] = []
        db_data_map[full_name].append(t)
        db_map[full_name] = t
        
        sorted_tokens = " ".join(sorted(full_name.split()))
        if sorted_tokens not in db_sorted_data_map: db_sorted_data_map[sorted_tokens] = []
        db_sorted_data_map[sorted_tokens].append(t)
        db_sorted_map[sorted_tokens] = t
        
    unique_db_names = list(set(db_names))
    
    matched = []
    sin_asignar = []
    conflictos = []

    def add_match_or_reject(x_name: str, x_norm: str, t) -> bool:
        if not t.dni or not t.dni.strip() or getattr(t, 'status', 'ACTIVO') != 'ACTIVO':
            sin_asignar.append({
                "id": t.id,
                "dni": t.dni or "",
                "apellidos": t.last_name or "",
                "nombres": t.first_name or "",
                "razon_social": t.razon_social or "",
                "normalized_name": t.normalized_name or x_norm,
                "status": "INCOMPLETO",
                "last_seen_at": t.last_seen_at.strftime("%Y-%m-%d") if getattr(t, 'last_seen_at', None) else "",
                "nombre_xml": x_name,
                "reason": "DATOS_INCOMPLETOS"
            })
            return False
        else:
            matched.append((x_name, t))
            return True
    
    # --- LOAD OVERRIDES ---
    # Performance: Preload all overrides to memory to avoid queries inside the loop O(n)
    all_overrides = db.query(TeacherNameOverride).all()
    
    scoped_overrides = {}
    global_overrides = {}
    
    latest_id = latest.id if latest else None
    
    for ov in all_overrides:
        if ov.xml_upload_id:
            if ov.xml_upload_id == latest_id:
                scoped_overrides[ov.xml_name_normalized] = ov
        else:
            global_overrides[ov.xml_name_normalized] = ov
            
    for x_name in xml_names:
        x_norm = normalize_name(x_name)
        
        # 0. Override Match Hierarchy (Priority 1: Scoped, Priority 2: Global)
        if x_norm in scoped_overrides:
            ov = scoped_overrides[x_norm]
            add_match_or_reject(x_name, x_norm, ov.teacher)
            logger.info(f"[MATCH ENGINE] '{x_name}' -> RESOLUTION_TYPE: SCOPED | TARGET_TEACHER_ID: {ov.teacher_id} | UPLOAD_ID: {latest_id}")
            continue
            
        if x_norm in global_overrides:
            ov = global_overrides[x_norm]
            add_match_or_reject(x_name, x_norm, ov.teacher)
            logger.info(f"[MATCH ENGINE] '{x_name}' -> RESOLUTION_TYPE: GLOBAL | TARGET_TEACHER_ID: {ov.teacher_id} | UPLOAD_ID: {latest_id}")
            continue
            
        logger.debug(f"[MATCH ENGINE] '{x_name}' -> RESOLUTION_TYPE: ALGO")
        
        x_sorted = " ".join(sorted(x_name.split()))
        
        # 1. Exact Match con múltiples registros en DB (Inconsistencia) -> CONFLICTO
        if x_name in db_data_map and len(db_data_map[x_name]) > 1:
            conflictos.append({
                "nombre_xml": x_name,
                "motivo": "MULTIPLES COINCIDENCIAS EXACTAS",
                "posibles_coincidencias": [{"teacher_id": str(t.id), "name": f"{t.last_name}, {t.first_name}", "dni": t.dni, "score": 1.0} for t in db_data_map[x_name]],
                "similitud": 1.0
            })
            continue
            
        # 1.5. Swapped Match con múltiples registros -> CONFLICTO
        if x_sorted in db_sorted_data_map and len(db_sorted_data_map[x_sorted]) > 1:
            conflictos.append({
                "nombre_xml": x_name,
                "motivo": "MULTIPLES COINCIDENCIAS INVERTIDAS",
                "posibles_coincidencias": [{"teacher_id": str(t.id), "name": f"{t.last_name}, {t.first_name}", "dni": t.dni, "score": 1.0} for t in db_sorted_data_map[x_sorted]],
                "similitud": 1.0
            })
            continue
            
        # 2. Exact Match (Strict Check) -> MATCHED
        if x_name in db_map:
            db_n = normalize_name(f"{db_map[x_name].last_name} {db_map[x_name].first_name}")
            if check_strict_match(x_name, db_n):
                add_match_or_reject(x_name, x_norm, db_map[x_name])
                continue
                
        # 3. Swapped Name Match (Strict Check) -> MATCHED
        if x_sorted in db_sorted_map:
            db_t = db_sorted_map[x_sorted]
            db_n = normalize_name(f"{db_t.last_name} {db_t.first_name}")
            if check_strict_match(x_name, db_n):
                add_match_or_reject(x_name, x_norm, db_t)
                continue
                
        # 4. Fuzzy Match
        best_matches = []
        is_matched = False
        
        for db_n in unique_db_names:
            score_normal = difflib.SequenceMatcher(None, x_name, db_n).ratio()
            db_sorted = " ".join(sorted(db_n.split()))
            score_sorted = difflib.SequenceMatcher(None, x_sorted, db_sorted).ratio()
            
            score = max(score_normal, score_sorted)
            
            if score >= 0.9:
                if check_strict_match(x_name, db_n):
                    add_match_or_reject(x_name, x_norm, db_map[db_n])
                    is_matched = True
                    break
                else:
                    best_matches.append((db_n, score))
            elif score >= 0.7:
                best_matches.append((db_n, score))
                
        if is_matched:
            continue
            
        if best_matches:
            best_matches.sort(key=lambda x: x[1], reverse=True)
            conflictos.append({
                "nombre_xml": x_name,
                "motivo": "SIMILITUD DUDOSA O RECHAZADO POR REGLA ESTRICTA",
                "posibles_coincidencias": [{"teacher_id": str(db_map[m[0]].id), "name": f"{db_map[m[0]].last_name}, {db_map[m[0]].first_name}", "dni": db_map[m[0]].dni, "score": round(m[1], 2)} for m in best_matches[:3]],
                "similitud": round(best_matches[0][1], 2)
            })
        else:
            sin_asignar.append({
                "id": None,
                "dni": "",
                "apellidos": "",
                "nombres": "",
                "razon_social": "",
                "normalized_name": x_norm,
                "status": "INCOMPLETO",
                "last_seen_at": "",
                "nombre_xml": x_name,
                "reason": "NO_EXISTE_EN_BD"
            })

    # MANDATORY LOGGING
    logger.info(f"TOTAL XML: {len(xml_names)}")
    logger.info(f"MATCHED: {len(matched)}")
    logger.info(f"CONFLICTOS: {len(conflictos)}")
    logger.info(f"SIN_ASIGNAR: {len(sin_asignar)}")
    
    if matched:
        logger.info(f"Top 5 MATCHED pairs: {[(m[0], normalize_name(f'{m[1].last_name} {m[1].first_name}')) for m in matched[:5]]}")
    if sin_asignar:
        logger.info(f"Top 5 SIN ASIGNAR: {[s['nombre_xml'] for s in sin_asignar[:5]]}")
    if conflictos:
        logger.info(f"Top 5 CONFLICTOS: {[(c['nombre_xml'], c['posibles_coincidencias']) for c in conflictos[:5]]}")
        
    # STEP 3 — SYNC STATUS AFTER CROSSING ENGINE
    conflict_teacher_ids = set()
    for conf in conflictos:
        for pos in conf.get("posibles_coincidencias", []):
            tid = pos.get("teacher_id")
            if tid:
                conflict_teacher_ids.add(str(tid))

    sin_asignar_teacher_ids = {str(item["id"]) for item in sin_asignar if item.get("id")}

    for teacher in db_teachers:
        t_id_str = str(teacher.id)
        if t_id_str in conflict_teacher_ids:
            apply_crossing_result(teacher, "CONFLICTO")
        elif t_id_str in sin_asignar_teacher_ids:
            apply_crossing_result(teacher, "SIN_ASIGNAR")
        else:
            apply_crossing_result(teacher, "MATCHED")

    db.flush()

    return {"matched": matched, "sin_asignar": sin_asignar, "conflictos": conflictos}

def get_sinasignar_crossed(db: Session, page: int = 1, limit: int = 20):
    """
    SIN ASIGNAR (Real Crossing): Nombres en XML que NO están en la DB, o en DB con estado INCOMPLETO.
    """
    results = _run_crossing_engine(db)
    sin_asignar_raw = list(results["sin_asignar"])
    conflictos = list(results["conflictos"])
    
    def is_mock_name(name_str: str) -> bool:
        if not name_str:
            return False
        u_name = name_str.upper()
        blacklist = ["TEST", "PRUEBA", "DEMO", "SANNCHEZ", "MARYA", "MOCK"]
        return any(b in u_name for b in blacklist)

    # Collect all unassigned items from XML and DB
    all_items = []
    for item in sin_asignar_raw:
        n_xml = item.get("nombre_xml") or ""
        ap = item.get("apellidos") or ""
        nom = item.get("nombres") or ""
        norm_name = item.get("normalized_name") or ""
        if is_mock_name(n_xml) or is_mock_name(f"{ap} {nom}") or is_mock_name(norm_name):
            continue
        all_items.append(item)
        
    existing_ids = {item["id"] for item in all_items if item.get("id")}
    
    latest = get_latest_completed_upload(db)
    xml_names = get_docente_names_from_xml(latest) if latest else []
    xml_norms = {normalize_name(name) for name in xml_names if name}
    
    print(f"[SIN_ASIGNAR ACTIVE XML]\nxml_names_count: {len(xml_names)}")
    logger.info(f"[SIN_ASIGNAR ACTIVE XML] xml_names_count: {len(xml_names)}")
    
    db_incompletes = db.query(Teacher).filter(
        Teacher.status == "INCOMPLETO",
        Teacher.merged_into_id.is_(None)
    ).all()
    
    for t in db_incompletes:
        t_norm = normalize_name(f"{t.last_name} {t.first_name}")
        if t_norm in xml_norms and t.id not in existing_ids:
            ap = t.last_name or ""
            nom = t.first_name or ""
            norm_name = t.normalized_name or ""
            n_xml = f"{ap}, {nom}".strip(", ")
            if is_mock_name(n_xml) or is_mock_name(f"{ap} {nom}") or is_mock_name(norm_name):
                continue
            all_items.append({
                "id": t.id,
                "dni": t.dni or "",
                "apellidos": ap,
                "nombres": nom,
                "razon_social": t.razon_social or "",
                "normalized_name": norm_name,
                "status": "INCOMPLETO",
                "last_seen_at": t.last_seen_at.strftime("%Y-%m-%d") if getattr(t, 'last_seen_at', None) else "",
                "nombre_xml": n_xml,
                "reason": "DATOS_INCOMPLETOS"
            })

            
    # Group items by normalized_name to detect collisions
    from collections import defaultdict
    groups = defaultdict(list)
    for item in all_items:
        norm = (item.get("normalized_name") or "").strip().upper()
        groups[norm].append(item)
        
    sin_asignar = []
    for norm, items in groups.items():
        if len(items) > 1:
            # Treat duplicate normalized names as CONFLICTO and add to conflictos
            for item in items:
                conflictos.append({
                    "nombre_xml": item.get("nombre_xml") or f"{item.get('apellidos', '')}, {item.get('nombres', '')}".strip(", "),
                    "motivo": "COLISION DE NOMBRE DUPLICADO EN SIN ASIGNAR",
                    "posibles_coincidencias": [{
                        "teacher_id": str(item["id"]) if item.get("id") else None, 
                        "name": f"{item.get('apellidos', '')}, {item.get('nombres', '')}".strip(", "), 
                        "dni": item.get("dni", ""), 
                        "score": 1.0
                    }],
                    "similitud": 1.0
                })
        else:
            sin_asignar.append(items[0])
            
    print(f"[SIN_ASIGNAR FILTERED]\ntotal_after_cleanup: {len(sin_asignar)}")
    logger.info(f"[SIN_ASIGNAR FILTERED] total_after_cleanup: {len(sin_asignar)}")
            
    logger.info("SIN_ASIGNAR SAMPLE", extra={"items": sin_asignar[:5]})
    logger.info("CONFLICTOS SAMPLE", extra={"items": conflictos[:5]})
    
    total = len(sin_asignar)
    start = (page - 1) * limit
    end = start + limit
    paged_items = sin_asignar[start:end]

    return StandardResponse(
        success=True,
        data=PaginatedResponseData(
            data=paged_items,
            total=total,
            page=page,
            limit=limit,
            total_pages=max(1, (total + limit - 1) // limit)
        )
    )

def get_conflictos_crossed(db: Session, page: int = 1, limit: int = 20):
    """
    CONFLICTOS (Real Crossing): Nombres en XML con múltiples coincidencias o similitud dudosa.
    """
    results = _run_crossing_engine(db)
    conflictos = results["conflictos"]
    
    total = len(conflictos)
    start = (page - 1) * limit
    end = start + limit
    paged_items = conflictos[start:end]

    return StandardResponse(
        success=True,
        data=PaginatedResponseData(
            data=paged_items,
            total=total,
            page=page,
            limit=limit,
            total_pages=max(1, (total + limit - 1) // limit)
        )
    )


def _get_fuzzy_match(db: Session, norm: str, original_name: str = "N/A", search_match_key: Optional[str] = None) -> Dict[str, Any]:
    """
    Motor de búsqueda endurecido (v4.0):
    1. Filtro por prefijo y longitud sobre 'normalized_for_match'.
    2. Regla de Protección: Comparación separada Apellido vs Nombre en forma canónica.
    3. Retorno estructurado con auditoría JSON en stdout.
    """
    default_res = {"decision": "NO_MATCH", "match_id": None, "score": 0, "candidate_name": None}
    if not norm or len(norm) < 2: return default_res
    
    from app.core.context import request_id_ctx
    req_id = request_id_ctx.get()

    # Si no nos pasan la match_key, la calculamos (para compatibilidad)
    if not search_match_key:
        search_match_key = normalize_teacher_name(norm, "")["match"]
    
    match_key = search_match_key
    min_l = int(len(match_key) * 0.8)
    max_l = int(len(match_key) * 1.2)
    
    # fetch_fuzzy_candidates retorna List[Tuple[UUID, str, str]] -> (id, canonical, match_key)
    candidates = repository.fetch_fuzzy_candidates(db, match_key, min_l, max_l)
    
    if not candidates:
        ev = {
            "event": "teacher_matching",
            "xml_name": original_name,
            "normalized_name": norm,
            "candidate": None,
            "score": 0,
            "decision": "NO_MATCH",
            "request_id": req_id,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
        print(json.dumps(ev))
        log_mdm_evidence(ev)
        return default_res

    best_candidate_id = None
    best_candidate_name = None
    max_score = 0
    final_decision = "NO_MATCH"

    # Preparar componentes del nombre de búsqueda (Canonical)
    search_parts = norm.split(' ', 1)
    s_ln = search_parts[0]
    s_fn = search_parts[1] if len(search_parts) > 1 else ""

    FUZZY_THRESHOLD = settings.FUZZY_THRESHOLD # 90-95 generalmente

    for c_id, cand_norm, cand_match_key in candidates:
        # Comparación global sobre Match Key (sin partículas, tokens ordenados)
        global_score = _fuzzy_ratio(match_key, cand_match_key)
        
        # Regla de Protección Estricta sobre Canonical
        is_hardened_dudoso = False
        cand_parts = cand_norm.split(' ', 1)
        c_ln = cand_parts[0]
        c_fn = cand_parts[1] if len(cand_parts) > 1 else ""
        
        # Protección: Si apellidos coinciden pero nombres no
        if len(s_fn) > 3 and s_ln and c_ln:
            ln_sim = _fuzzy_ratio(s_ln, c_ln) / 100.0
            fn_sim = _fuzzy_ratio(s_fn, c_fn) / 100.0
            if ln_sim > 0.95 and fn_sim < 0.70:
                is_hardened_dudoso = True

        # Clasificación de decisión
        current_decision = "NO_MATCH"
        if global_score >= FUZZY_THRESHOLD and not is_hardened_dudoso:
            current_decision = "MATCH_AUTOMATICO"
        elif global_score >= 75 or is_hardened_dudoso:
            current_decision = "MATCH_DUDOSO"
        
        # Lógica de Selección del Mejor Candidato
        if current_decision == "MATCH_AUTOMATICO":
            if final_decision != "MATCH_AUTOMATICO" or global_score > max_score:
                final_decision = "MATCH_AUTOMATICO"
                max_score = global_score
                best_candidate_id = c_id
                best_candidate_name = cand_norm
        elif current_decision == "MATCH_DUDOSO" and final_decision != "MATCH_AUTOMATICO":
            if global_score > max_score or final_decision == "NO_MATCH":
                final_decision = "MATCH_DUDOSO"
                max_score = global_score
                best_candidate_id = c_id
                best_candidate_name = cand_norm

    res = {
        "decision": final_decision,
        "match_id": best_candidate_id,
        "score": max_score,
        "candidate_name": best_candidate_name
    }
    
    # Auditoría Final JSON
    ev = {
        "event": "teacher_matching",
        "xml_name": original_name,
        "normalized_name": norm,
        "candidate": best_candidate_name,
        "score": max_score,
        "decision": final_decision,
        "request_id": req_id,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    print(json.dumps(ev))
    log_mdm_evidence(ev)
    
    return res

    # NO_MATCH
    ev = {
        "event": "teacher_matching",
        "xml_name": original_name,
        "normalized_name": norm,
        "candidate": None,
        "score": 0,
        "decision": "NO_MATCH"
    }
    logger.info(f"MDM: No se encontró match para {original_name}", extra=ev)
    log_mdm_evidence(ev)
    return default_res


def cross_check_xml_teachers(db: Session, teachers_xml: List[Dict[str, str]]) -> Dict[str, Any]:
    """
    Compara docentes del XML contra la maestra.
    No encontrados → teachers_sinasignar (sin duplicar por normalized_name).
    """
    now_utc = datetime.now(timezone.utc)
    nuevos = ya_maestra = ya_sinasignar = 0

    for t in teachers_xml:
        fn = (t.get("first_name") or "").strip()
        ln = (t.get("last_name") or "").strip()
        if not fn and not ln:
            continue

        norm_dict = normalize_teacher_name(ln, fn)
        norm = norm_dict["canonical"]
        match_key = normalize_for_match(f"{ln} {fn}")


        # 1. Match Exacto en Maestra
        if repository.fetch_teacher_by_normalized(db, norm):
            ya_maestra += 1
            continue

        # 2. Match Exacto en SinAsignar
        existing_sa = repository.fetch_sinasignar_by_normalized(db, norm)
        if existing_sa:
            repository.increment_sinasignar_detection(db, existing_sa)
            ya_sinasignar += 1
            continue

        # 3. Match FUZZY optimizado (v3.2)
        possible_match_id = _get_fuzzy_match(db, norm, f"{ln}, {fn}")
        is_possible_duplicate = possible_match_id is not None

        repository.create_sinasignar(db, {
            "last_name":       ln,
            "first_name":      fn,
            "normalized_name": norm,
            "normalized_for_match": match_key,
            "source":          "xml",
            "times_detected":  1,
            "last_seen_at":    now_utc,
            "possible_duplicate": is_possible_duplicate,
            "possible_match_id":  possible_match_id,

        })
        nuevos += 1

    db.commit()
    return {
        "success":          True,
        "nuevos_sinasignar": nuevos,
        "ya_en_maestra":    ya_maestra,
        "ya_en_sinasignar": ya_sinasignar,
        "detalle_nuevos":   [],
    }


# ════════════════════════════════════════════════════════
#  REPROCESAMIENTO HISTÓRICO — NUEVO ENDPOINT
# ════════════════════════════════════════════════════════

FUZZY_THRESHOLD = 90.0


def reprocesar_historico(db: Session) -> Dict[str, Any]:
    """
    1. Lee todos los Teacher de la BD (provienen de XMLs históricos).
    2. Puebla normalized_name donde falta.
    3. Obtiene el set de normalized_name de la maestra (con DNI o enriched).
    4. Para cada teacher no en maestra: fuzzy match. Si ≥ 90 → conflicto.
    5. Si no hay match: inserta en sinasignar (o incrementa times_detected).
    Retorna estadísticas y lista de conflictos fuzzy para resolución humana.
    """
    logger.info("[reprocesar_historico] Iniciando...")
    now_utc = datetime.now(timezone.utc)

    # ── Paso 1: Cargar todos los teachers del sistema ──────────
    all_teachers: List[Teacher] = repository.fetch_all_teachers_from_table(db)
    logger.info(f"  Teachers en BD: {len(all_teachers)}")

    # ── Paso 2: Poblar normalized_name donde esté vacío ───────
    updates_needed: List[Tuple[Any, str]] = []
    for t in all_teachers:
        if not t.normalized_name:
            norm = normalize_teacher_name(t.last_name or "", t.first_name or "")
            if norm:
                norm_str = ensure_string(norm)
                updates_needed.append((t.id, norm_str))
                t.normalized_name = norm_str  # local cache

    if updates_needed:
        repository.bulk_update_normalized(db, updates_needed)
        db.flush()
        logger.info(f"  Normalized_name poblado para {len(updates_needed)} registros")

    # ── Paso 3: Construir set de maestra (teachers con normalized) ─
    maestra_set = {
        t.normalized_name
        for t in all_teachers
        if t.normalized_name
    }
    maestra_list = sorted(maestra_set)  # para fuzzy

    # ── Paso 4: Construir set de docentes no asignados existentes ──
    sinasignar_norms = {
        s.normalized_name
        for s in db.query(Teacher).filter(Teacher.is_assigned == False).all()
    }

    # ── Paso 5: Analizar cada teacher ─────────────────────────
    analizados = 0
    nuevos = 0
    duplicados = 0
    conflictos: List[Dict[str, Any]] = []

    for t in all_teachers:
        norm = t.normalized_name
        if not norm:
            continue

        analizados += 1

        # Ya está en maestra (debería siempre ser cierto ya que viene de allí)
        # Buscamos si el teacher tiene match por parte del Excel (dni / razon_social)
        # La lógica real: teachers que NO tienen dni ni razon_social son "solo XML"
        is_xml_only = (not t.dni and not t.razon_social)
        if not is_xml_only:
            continue  # Ya enriquecido por Excel — no procesar

        # ── Nivel 1: Exact match en maestra enriched ───────────
        # Buscar si existe otro teacher (con dni/razon_social) con mismo normalized_name
        maestra_enriched = {
            t2.normalized_name
            for t2 in all_teachers
            if (t2.dni or t2.razon_social) and t2.normalized_name
        }
        if norm in maestra_enriched:
            logger.debug(f"  SKIP (exact maestra enriched): {norm}")
            continue

        # ── Nivel 2: Fuzzy match ───────────────────────────────
        best_match_norm: Optional[str] = None
        best_score: float = 0.0

        for m_norm in maestra_enriched:
            score = _fuzzy_ratio(norm, m_norm)
            if score > best_score:
                best_score = score
        # B) Match Fuzzy optimizado
        possible_match_id = _get_fuzzy_match(db, norm)
        if possible_match_id:
            match_t = repository.fetch_teacher_by_id_full(db, possible_match_id)
            conflictos.append({
                "teacher_id": t.id,
                "detected": f"{t.last_name}, {t.first_name}",
                "detected_norm": norm,
                "suggested_norm": match_t.normalized_name if match_t else "ID:" + str(possible_match_id),
                "score": _fuzzy_ratio(norm, match_t.normalized_name) if match_t else 0
            })
            continue

        # C) Sin match -> Crear en SinAsignar o actualizar
        if norm in sinasignar_norms:
            # Ya existe → incrementar contador
            existing_sa = repository.fetch_sinasignar_by_normalized(db, norm)
            if existing_sa:
                repository.increment_sinasignar_detection(db, existing_sa)
            duplicados += 1
            logger.debug(f"  DUPLICATE sinasignar (incrementado): {norm}")
        else:
            repository.create_sinasignar(db, {
                "apellidos":       t.last_name or "",
                "nombres":         t.first_name or "",
                "normalized_name": norm,
                "source":          "xml",
                "times_detected":  1,
                "last_seen_at":    now_utc,
            })
            sinasignar_norms.add(norm)
            nuevos += 1
            logger.info(f"  NUEVO sinasignar: {norm}")

    db.commit()

    logger.info(
        f"[reprocesar_historico] Fin — "
        f"analizados={analizados}, nuevos={nuevos}, "
        f"duplicados={duplicados}, fuzzy={len(conflictos)}"
    )
    return {
        "success":                    True,
        "total_docentes_analizados":  analizados,
        "nuevos_insertados":          nuevos,
        "duplicados_detectados":      duplicados,
        "posibles_coincidencias_fuzzy": conflictos,
    }


# ════════════════════════════════════════════════════════
#  RESOLUCIÓN DE CONFLICTOS FUZZY
# ════════════════════════════════════════════════════════

def vincular_teacher_alias(db: Session, teacher_id: UUID, target_norm: str) -> Dict[str, Any]:
    """
    'Vincular': actualiza el normalized_name del teacher para que coincida
    con el target (teacher enriquecido por maestra). Esto resuelve la ambigüedad.
    """
    t = repository.fetch_teacher_by_id_full(db, teacher_id)
    if not t:
        raise ValueError("Docente no encontrado")

    old_norm = t.normalized_name
    repository.update_teacher(db, t, {"normalized_name": target_norm})

    # Si estaba en sinasignar con el normalized_name viejo, eliminarlo
    sa = repository.fetch_sinasignar_by_normalized(db, old_norm)
    if sa:
        repository.delete_sinasignar(db, sa)

    # 3. Sincronizar incidencias que usaban el alias/nombre viejo
    synced = obs_repo.sync_teacher_id_by_name(db, teacher_id, old_norm)
    
    db.commit()
    msg = f"Alias vinculado: '{old_norm}' → '{target_norm}'."
    if synced > 0:
        msg += f" {synced} incidencias actualizadas."
    return {"success": True, "message": msg}


# ════════════════════════════════════════════════════════
#  CRUD — TEACHERS_SINASIGNAR
# ════════════════════════════════════════════════════════

def _sinasignar_to_dict(obj: Teacher) -> Dict[str, Any]:
    """Usa el modelo Teacher unificado para representar registros Sin Asignar."""
    return {
        "id":             str(obj.id),
        "dni":            obj.dni or "",
        "apellidos":      obj.last_name,
        "nombres":        obj.first_name,
        "razon_social":   getattr(obj, 'razon_social', ""),
        "normalized_name": obj.normalized_name,
        "source":         obj.source or "xml",
        "times_detected": obj.times_detected or 1,
        "is_possible_duplicate": obj.possible_duplicate,
        "last_seen_at":   obj.last_seen_at.strftime("%Y-%m-%d") if obj.last_seen_at else "",
    }


def get_sinasignar(db: Session, page: int = 1, limit: int = 20):
    # REGLA v5: Solo los que están en estado INCOMPLETO
    q = db.query(Teacher).filter(Teacher.status == "INCOMPLETO").order_by(
        Teacher.last_name,
        Teacher.first_name
    )
    total = q.count()
    items = q.offset((page - 1) * limit).limit(limit).all()

    data_items = [_sinasignar_to_dict(o) for o in items]
    
    return StandardResponse(
        success=True,
        data=PaginatedResponseData(
            data=data_items,
            total=total,
            page=page,
            limit=limit,
            total_pages=max(1, (total + limit - 1) // limit)
        )
    )


def update_sinasignar_item(db: Session, sid: UUID, payload: Dict[str, Any]) -> Dict[str, Any]:
    obj = repository.fetch_sinasignar_by_id(db, sid)
    if not obj:
        raise ValueError("Registro no encontrado")

    apellidos = (payload.get("apellidos") or obj.apellidos).strip()
    nombres   = (payload.get("nombres") or obj.nombres).strip()
    norm      = normalize_teacher_name(apellidos, nombres)

    repository.update_sinasignar(db, obj, {
        "dni":             (payload.get("dni") or "").strip() or None,
        "last_name":       apellidos,
        "first_name":      nombres,
        "razon_social":    (payload.get("razon_social") or "").strip() or None,
        "normalized_name": norm,
    })
    db.commit()
    return _sinasignar_to_dict(obj)


def promote_sinasignar(db: Session, identifier: UUID) -> Dict[str, Any]:
    # REGLA FASE 2: Buscar en tabla única
    obj = resolve_teacher(db, identifier)
    if not obj:
        raise ValueError("Registro no encontrado")

    if obj.is_assigned:
        return {"action": "already_promoted", "message": "Este docente ya está asignado.", "teacher": _teacher_to_dict(obj)}

    # Verificar si ya existe un docente titular con el mismo nombre (Fusión implícita)
    existing = db.query(Teacher).filter(
        Teacher.normalized_name == obj.normalized_name,
        Teacher.is_assigned == True,
        Teacher.uid != obj.uid
    ).first()

    if existing:
        # Transferimos relaciones de este 'sin asignar' al titular y borramos el duplicado
        repository.merge_teachers_db(db, existing.id, obj.id)
        db.commit()
        return {"action": "merged", "message": f"Se fusionó con el docente titular existente (id={existing.id}).", "teacher": _teacher_to_dict(existing)}

    # Promoción simple
    obj.is_assigned = True
    obj.source = f"PROMOTED_{int(time_lib.time())}"
    
    # Sincronizar incidencias
    obs_repo.sync_teacher_id_by_name(db, obj.id, obj.normalized_name)
    
    db.commit()
    return {"action": "promoted", "message": "Docente promovido exitosamente", "teacher": _teacher_to_dict(obj)}


def delete_sinasignar_item(db: Session, identifier: UUID) -> bool:
    obj = resolve_teacher(db, identifier)
    if not obj:
        raise ValueError("Registro no encontrado")
    repository.delete_sinasignar(db, obj)
    db.commit()
    return True


# ════════════════════════════════════════════════════════
#  FUSIÓN — MERGE
# ════════════════════════════════════════════════════════

def merge_teachers(db: Session, main_id: UUID, merge_id: UUID) -> Dict[str, Any]:
    """
    Orquestación de la fusión de dos docentes.
    REGLA: Debe ser atómica y recalcular normalized_name.
    """
    if main_id == merge_id:
        raise ValueError("No se puede fusionar un docente con el mismo registro.")

    t_main = resolve_teacher(db, main_id)
    t_merge = resolve_teacher(db, merge_id)

    if not t_main or not t_merge:
        raise ValueError("Uno o ambos docentes no existen.")

    try:
        # 1. Ejecutar transferencia de relaciones y eliminación del secundario
        secondary_norm_copy = t_merge.normalized_name # Guardamos para limpieza
        secondary_id_copy   = t_merge.id
        
        repository.merge_teachers_db(db, main_id, merge_id)

        # 2. Recalcular normalized_name del resultante para asegurar consistencia
        new_norm = normalize_teacher_name(t_main.last_name or "", t_main.first_name or "")
        repository.update_teacher(db, t_main, {"normalized_name": new_norm})

        # 3. Limpieza DIRIGIDA (v3.2) de SinAsignar
        repository.cleanup_sinasignar_post_merge(db, secondary_id_copy, secondary_norm_copy)
        # También limpiar el nombre nuevo del principal si existía en pendientes
        repository.cleanup_sinasignar_post_merge(db, main_id, new_norm)

        # 4. Sincronizar incidencias del eliminado (por nombre) al nuevo ID principal
        # Especialmente útil si había incidencias con el nombre del secundario pero sin ID
        synced_sec = obs_repo.sync_teacher_id_by_name(db, main_id, secondary_norm_copy)
        
        db.commit()
        logger.info(f"FUSIÓN EXITOSA: main_id={main_id}, merge_id={merge_id}. {synced_sec} incidencias huérfanas vinculadas.")
        return {
            "success": True,
            "message": f"Fusión completada. {synced_sec} incidencias reconectadas.",
            "teacher": _teacher_to_dict(t_main)
        }
    except Exception as e:
        db.rollback()
        logger.error(f"ERROR EN FUSIÓN: {str(e)}")
        raise RuntimeError(f"Error crítico en la fusión. Operación cancelada (Rollback ejecutado): {str(e)}")


def get_potential_duplicates(db: Session) -> List[Dict[str, Any]]:
    """Obtiene docentes marcados como posibles duplicados para auditoría (v3.2)."""
    duplicates = db.query(Teacher).filter(Teacher.possible_duplicate == True).all()
    # Usar active_ids vacío como fallback para el mapeo simple de auditoría
    return [_teacher_to_dict(t, set()) for t in duplicates]


def get_matching_preview(db: Session) -> List[Dict[str, Any]]:
    """
    Diagnóstico de Matching (v4.5):
    Compara XML (Sin Asignar) vs Maestra (is_assigned=True) basándose en normalized_for_match.
    """
    xml_teachers = db.query(Teacher).filter(Teacher.is_assigned == False).all()
    maestra = db.query(Teacher).filter(Teacher.is_assigned == True, Teacher.merged_into_id == None).all()
    
    # Mapeo de maestra por normalized_for_match para búsqueda rápida
    maestra_map = {}
    for t in maestra:
        nm = t.normalized_for_match or ""
        if not nm: continue
        if nm not in maestra_map:
            maestra_map[nm] = []
        maestra_map[nm].append(t)
        
    results = []
    for xt in xml_teachers:
        nm = xt.normalized_for_match or ""
        matches = maestra_map.get(nm, [])
        
        status = "NO_MATCH"
        if len(matches) == 1:
            status = "MATCH_FOUND"
        elif len(matches) > 1:
            status = "MULTIPLE_MATCH"
            
        results.append({
            "xml_teacher_id": str(xt.id),
            "xml_name": f"{xt.last_name}, {xt.first_name}",
            "normalized_for_match": nm,
            "status": status,
            "candidates": [
                {
                    "id": str(m.id), 
                    "name": f"{m.last_name}, {m.first_name}",
                    "dni": m.dni
                } for m in matches
            ]
        })
        
    return results

def resolve_conflict(db: Session, xml_name_raw: str, teacher_id: UUID, is_global: bool = False, current_user_id: Optional[UUID] = None) -> UUID:
    """
    Registra un override (Scoped o Global) y reasigna para el engine_crossing_run.
    """
    from sqlalchemy.exc import IntegrityError
    
    xml_name_normalized = normalize_name(xml_name_raw)
    
    latest_upload_id = None
    if not is_global:
        latest = get_latest_completed_upload(db)
        if not latest:
            raise ValueError("No se puede crear un override local porque no hay uploads completados.")
        latest_upload_id = latest.id

    # Check existence to update or insert
    existing = db.query(TeacherNameOverride).filter(
        TeacherNameOverride.xml_name_normalized == xml_name_normalized,
        TeacherNameOverride.xml_upload_id == latest_upload_id
    ).first()

    if existing:
        existing.teacher_id = teacher_id
        existing.confidence = 1.0
        override_id = existing.id
    else:
        new_override = TeacherNameOverride(
            xml_name_raw=xml_name_raw,
            xml_name_normalized=xml_name_normalized,
            teacher_id=teacher_id,
            xml_upload_id=latest_upload_id,
            confidence=1.0
        )
        db.add(new_override)
        db.flush()
        override_id = new_override.id
        
    db.commit()
    logger.info(f"[RESOLVE CONFLICT] {'GLOBAL' if is_global else 'SCOPED'} override created for '{xml_name_normalized}' -> TeacherID: {teacher_id}")
    return override_id

def undo_conflict(db: Session, override_id: UUID):
    """
    Elimina un override previamente registrado.
    """
    override = db.query(TeacherNameOverride).filter(TeacherNameOverride.id == override_id).first()
    if not override:
        raise ValueError("Override no encontrado")
        
    db.delete(override)
    db.commit()
    logger.info(f"[UNDO CONFLICT] Override {override_id} deleted.")
