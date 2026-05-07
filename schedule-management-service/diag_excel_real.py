"""
Diagnostico del Excel REAL del usuario.
Busca cualquier .xlsx en el directorio y lo inspecciona.
"""
import sys, io, os, glob
sys.path.insert(0, '.')

import openpyxl

# Buscar Excels en Desktop y Downloads
search_paths = [
    r"C:\Users\SISTEMAS2\Desktop\*.xlsx",
    r"C:\Users\SISTEMAS2\Desktop\*.xls",
    r"C:\Users\SISTEMAS2\Downloads\*.xlsx",
    r"C:\Users\SISTEMAS2\Downloads\*.xls",
    r"D:\Desktop\*.xlsx",
    r"D:\Desktop\*.xls",
    r"D:\Desktop\MOD HOR\*.xlsx",
    r"D:\Desktop\MOD HOR\*.xls",
]

found = []
for pattern in search_paths:
    found.extend(glob.glob(pattern))

if not found:
    print("[DIAG] No se encontraron archivos Excel en ubicaciones conocidas")
    print("[DIAG] Buscando en todo el Desktop...")
    for root, dirs, files in os.walk(r"D:\Desktop"):
        dirs[:] = [d for d in dirs if d not in ['__pycache__', '.git', 'node_modules']]
        for f in files:
            if f.lower().endswith(('.xlsx', '.xls')):
                found.append(os.path.join(root, f))
                if len(found) >= 10:
                    break
        if len(found) >= 10:
            break

print("[DIAG] Archivos Excel encontrados:")
for f in found:
    size = os.path.getsize(f)
    print("  %s (%d bytes)" % (f, size))

if not found:
    print("[DIAG] NINGUNO encontrado. El usuario debe subir el archivo manualmente.")
    sys.exit(0)

# Inspeccionar el primero que parezca de docentes
target = None
for f in found:
    name_lower = f.lower()
    if any(kw in name_lower for kw in ['docente', 'teacher', 'maestra', 'planilla', 'personal']):
        target = f
        break

if not target:
    target = found[0]

print("\n[DIAG] Inspeccionando: %s" % target)

wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
print("[DIAG] Hojas: %s" % wb.sheetnames)
ws = wb.active
print("[DIAG] Hoja activa: %s" % ws.title)
print("[DIAG] max_row=%s max_column=%s" % (ws.max_row, ws.max_column))

# Cabeceras reales
headers_raw = []
for c in range(1, min((ws.max_column or 0) + 1, 20)):
    val = ws.cell(1, c).value
    headers_raw.append(str(val or "").strip())

print("[DIAG] Cabeceras EXACTAS (col 1..%d):" % len(headers_raw))
for i, h in enumerate(headers_raw, 1):
    print("  col%d: repr=%r  upper=%r" % (i, h, h.upper()))

# Primeras 5 filas de datos
print("\n[DIAG] Primeras 5 filas de datos:")
for row_idx in range(2, min(7, (ws.max_row or 2) + 1)):
    row_data = []
    for c in range(1, len(headers_raw) + 1):
        val = ws.cell(row_idx, c).value
        row_data.append(repr(val))
    print("  row%d: %s" % (row_idx, row_data))

# Validar col_map con ALIASES
ALIASES = {
    "DNI":          ["DNI", "RUC", "DOCUMENTO"],
    "APELLIDOS":    ["APELLIDOS", "APELLIDO", "LAST_NAME"],
    "NOMBRES":      ["NOMBRES", "NOMBRE", "FIRST_NAME"],
    "RAZON_SOCIAL": ["RAZON SOCIAL", "RAZON_SOCIAL", "EMPRESA", "RAZON SOCIAL"],
}
col_map = {}
upper_headers = [h.upper() for h in headers_raw]
for canonical, aliases in ALIASES.items():
    for idx, h in enumerate(upper_headers, start=1):
        if h in aliases:
            col_map[canonical] = idx
            break

print("\n[DIAG] col_map resuelto: %s" % col_map)
missing = [k for k in ["APELLIDOS", "NOMBRES"] if k not in col_map]
if missing:
    print("[DIAG] BUG ENCONTRADO: Faltan columnas obligatorias: %s" % missing)
    print("[DIAG] Las cabeceras del Excel no coinciden con los aliases esperados")
    print("[DIAG] Cabeceras reales (upper): %s" % upper_headers)
    print("[DIAG] Aliases buscados por APELLIDOS: %s" % ALIASES['APELLIDOS'])
else:
    print("[DIAG] OK: Columnas obligatorias mapeadas correctamente")
