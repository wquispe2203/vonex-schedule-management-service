"""
Diagnostico real de import_excel.
"""
import sys, io, os
sys.path.insert(0, '.')
os.environ["PYTHONIOENCODING"] = "utf-8"

import openpyxl
from app.core.database import SessionLocal
from app.models import Teacher
from sqlalchemy import func

def make_test_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["DNI", "APELLIDOS", "NOMBRES", "RAZON SOCIAL"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

TEST_ROWS = [
    ["12345678", "GARCIA LOPEZ",   "JUAN CARLOS",  "EMPRESA A SAC"],
    ["87654321", "PEREZ MARTINEZ", "MARIA ELENA",  "EMPRESA B SAC"],
    ["",         "TORRES QUISPE",  "PEDRO MANUEL", ""],
    ["",         "",               "",             ""],  # fila vacia
]

excel_bytes = make_test_excel(TEST_ROWS)
print("[DIAG] Excel creado: %d bytes" % len(excel_bytes))

# --- PARSING PURO ---
print("\n[DIAG] === PARSING PURO ===")
wb2 = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
ws2 = wb2.active
print("[DIAG] max_row=%d max_column=%d" % (ws2.max_row, ws2.max_column))

headers_raw = [str(ws2.cell(1, c).value or "").strip().upper() for c in range(1, ws2.max_column + 1)]
print("[DIAG] Headers: %s" % headers_raw)

ALIASES = {
    "DNI":          ["DNI", "RUC", "DOCUMENTO"],
    "APELLIDOS":    ["APELLIDOS", "APELLIDO", "LAST_NAME"],
    "NOMBRES":      ["NOMBRES", "NOMBRE", "FIRST_NAME"],
    "RAZON_SOCIAL": ["RAZON SOCIAL", "RAZON_SOCIAL", "EMPRESA"],
}
col_map = {}
for canonical, aliases in ALIASES.items():
    for idx, h in enumerate(headers_raw, start=1):
        if h in aliases:
            col_map[canonical] = idx
            break

print("[DIAG] col_map: %s" % col_map)

for row_idx in range(2, ws2.max_row + 1):
    def cell(key, ri=row_idx):
        idx = col_map.get(key)
        if idx is None: return ""
        val = ws2.cell(ri, idx).value
        return str(val).strip() if val is not None else ""

    ap = cell("APELLIDOS")
    nom = cell("NOMBRES")
    dni = cell("DNI") or None
    print("[DIAG] row=%d ap=[%s] nom=[%s] dni=[%s] -> %s" % (
        row_idx, ap, nom, dni, "SKIP" if not ap and not nom else "PROCESS"))

# --- PRIMERA PASADA (INSERT) ---
print("\n[DIAG] === PRIMERA PASADA (deberia insertar) ===")
db = SessionLocal()
try:
    before = db.query(func.count(Teacher.id)).scalar()
    print("[DIAG] Teachers antes: %d" % before)

    from app.modules.docentes.service import import_excel
    result = import_excel(db, excel_bytes)

    after = db.query(func.count(Teacher.id)).scalar()
    print("[DIAG] Teachers despues: %d (delta=%d)" % (after, after-before))
    print("[DIAG] inserted=%d updated=%d skipped=%d rows_count=%d" % (
        result['inserted'], result['updated'], result['skipped'], len(result['rows'])))

    for r in result['rows']:
        print("  ROW[%d] %s -> %s" % (r['row'], r['apellidos'], r['action']))

    if result['inserted'] == 0 and result['updated'] == 0:
        print("[DIAG] BUG CONFIRMADO: inserted=0 updated=0")
    else:
        print("[DIAG] OK: inserted=%d updated=%d" % (result['inserted'], result['updated']))
finally:
    db.close()

# --- SEGUNDA PASADA (UPDATE) ---
print("\n[DIAG] === SEGUNDA PASADA (deberia actualizar) ===")
db2 = SessionLocal()
try:
    result2 = import_excel(db2, excel_bytes)
    print("[DIAG] inserted=%d updated=%d" % (result2['inserted'], result2['updated']))
    if result2['inserted'] > 0:
        print("[DIAG] BUG: Segunda pasada inserto en lugar de actualizar")
    elif result2['updated'] > 0:
        print("[DIAG] OK: Segunda pasada actualizo correctamente")
    else:
        print("[DIAG] BUG: Segunda pasada tambien dio 0/0")
finally:
    db2.close()
